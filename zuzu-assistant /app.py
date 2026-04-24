import streamlit as st
import io
import os
import anthropic
import openpyxl
from google.oauth2.credentials import Credentials
from google.auth.transport.requests import Request
from googleapiclient.discovery import build

st.set_page_config(page_title="Zuzu Assistant", page_icon="🌸", layout="wide")
st.title("Zuzu Assistant")

def get_credentials():
    creds = Credentials(
        token=None,
        refresh_token=st.secrets["google"]["refresh_token"],
        client_id=st.secrets["google"]["client_id"],
        client_secret=st.secrets["google"]["client_secret"],
        token_uri="https://oauth2.googleapis.com/token",
        scopes=[
            "https://www.googleapis.com/auth/spreadsheets",
            "https://www.googleapis.com/auth/drive"
        ]
    )
    creds.refresh(Request())
    return creds

@st.cache_resource
def get_services():
    creds = get_credentials()
    sheets = build("sheets", "v4", credentials=creds)
    drive = build("drive", "v3", credentials=creds)
    return sheets, drive

def list_files(drive):
    results = drive.files().list(
        q="mimeType='application/vnd.google-apps.spreadsheet' or mimeType='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'",
        fields="files(id, name, mimeType)",
        corpora="user",
        pageSize=100
    ).execute()
    return results.get("files", [])

def read_google_sheet(sheets, file_id, tab=None):
    if tab is None:
        info = sheets.spreadsheets().get(spreadsheetId=file_id).execute()
        tab = info["sheets"][0]["properties"]["title"]
    result = sheets.spreadsheets().values().get(
        spreadsheetId=file_id, range=tab
    ).execute()
    return result.get("values", []), tab

def read_xlsx(drive, file_id):
    request = drive.files().get_media(fileId=file_id)
    data = io.BytesIO(request.execute())
    wb = openpyxl.load_workbook(data)
    return {
        name: [[str(c.value) if c.value is not None else "" for c in row]
               for row in wb[name].iter_rows()]
        for name in wb.sheetnames
    }

def write_google_sheet(sheets, file_id, tab, data):
    sheets.spreadsheets().values().update(
        spreadsheetId=file_id,
        range=tab,
        valueInputOption="USER_ENTERED",
        body={"values": data}
    ).execute()

def format_for_claude(data, max_rows=300):
    if not data:
        return "Empty sheet."
    return "\n".join(["\t".join(row) for row in data[:max_rows]])

def ask_claude(user_message, files, sheets, drive):
    client = anthropic.Anthropic(api_key=st.secrets["anthropic"]["api_key"])

    file_list = "\n".join([
        "- " + f["name"] + " (ID: " + f["id"] + ", type: " + ("Google Sheet" if "spreadsheet" in f["mimeType"] else "xlsx") + ")"
        for f in files
    ])

    system = """You are Zuzu Assistant, a smart planning assistant for Zoe McDaniel, founder of Zuzu Collective, a boutique wedding and event planning company in San Diego.

You have access to all of Zoe's planning spreadsheets. Here they are:

""" + file_list + """

Zoe's active clients:
- Berenice & Nello - Fullerton Community Center, April 18, 2026
- Becky & Russ - Raven's Roost, Fort Bragg, July 5, 2026
- Chris & Lindsay - South Coast Botanic Gardens, Sept 5, 2026
- Cody & Danny - Ace Hotel Palm Springs, Sept 26, 2026
- Natalie & Josh - November 15, 2026
- Ana & Nick - Orcas Island, July 17, 2027

When Zoe asks about priorities, clients, vendors, or timelines, read the relevant sheets and give specific actionable answers. When she asks you to make edits, use write_sheet and summarize what changed. Be direct and efficient."""

    tools = [
        {
            "name": "read_sheet",
            "description": "Read a planning spreadsheet from Google Drive",
            "input_schema": {
                "type": "object",
                "properties": {
                    "file_id": {"type": "string"},
                    "file_name": {"type": "string"},
                    "tab": {"type": "string"}
                },
                "required": ["file_id", "file_name"]
            }
        },
        {
            "name": "write_sheet",
            "description": "Write updated data to a Google Sheet",
            "input_schema": {
                "type": "object",
                "properties": {
                    "file_id": {"type": "string"},
                    "tab": {"type": "string"},
                    "data": {"type": "array", "items": {"type": "array"}},
                    "summary": {"type": "string"}
                },
                "required": ["file_id", "tab", "data", "summary"]
            }
        }
    ]

    messages = [{"role": "user", "content": user_message}]
    pending_writes = []

    while True:
        response = client.messages.create(
            model="claude-sonnet-4-6",
            max_tokens=4096,
            system=system,
            tools=tools,
            messages=messages
        )

        messages.append({"role": "assistant", "content": response.content})

        if response.stop_reason == "end_turn":
            text = next((b.text for b in response.content if hasattr(b, "text")), "")
            return text, pending_writes

        if response.stop_reason == "tool_use":
            tool_results = []
            for block in response.content:
                if block.type != "tool_use":
                    continue
                if block.name == "read_sheet":
                    fid = block.input["file_id"]
                    fname = block.input["file_name"]
                    tab = block.input.get("tab")
                    file_info = next((f for f in files if f["id"] == fid), None)
                    try:
                        if file_info and "openxmlformats" in file_info["mimeType"]:
                            sheets_data = read_xlsx(drive, fid)
                            tab = tab if tab in sheets_data else list(sheets_data.keys())[0]
                            data = sheets_data[tab]
                        else:
                            data, tab = read_google_sheet(sheets, fid, tab)
                        result = "Contents of " + fname + " (tab: " + tab + "):\n\n" + format_for_claude(data)
                    except Exception as e:
                        result = "Error reading " + fname + ": " + str(e)
                    tool_results.append({"type": "tool_result", "tool_use_id": block.id, "content": result})

                elif block.name == "write_sheet":
                    pending_writes.append({
                        "file_id": block.input["file_id"],
                        "tab": block.input["tab"],
                        "data": block.input["data"],
                        "summary": block.input["summary"]
                    })
                    tool_results.append({"type": "tool_result", "tool_use_id": block.id, "content": "Write queued: " + block.input["summary"]})

            messages.append({"role": "user", "content": tool_results})

try:
    sheets_svc, drive_svc = get_services()
    if "files" not in st.session_state:
        with st.spinner("Connecting to your Drive..."):
            st.session_state.files = list_files(drive_svc)
    files = st.session_state.files
    st.caption(str(len(files)) + " planning files connected")
except Exception as e:
    st.error("Could not connect to Google Drive: " + str(e))
    st.stop()

if "messages" not in st.session_state:
    st.session_state.messages = []
if "pending_writes" not in st.session_state:
    st.session_state.pending_writes = []

for msg in st.session_state.messages:
    with st.chat_message(msg["role"]):
        st.markdown(msg["content"])

if st.session_state.pending_writes:
    for i, write in enumerate(st.session_state.pending_writes):
        st.warning("Proposed edit: " + write["summary"])
        col1, col2 = st.columns([1, 4])
        if col1.button("Confirm", key="confirm_" + str(i)):
            write_google_sheet(sheets_svc, write["file_id"], write["tab"], write["data"])
            st.session_state.pending_writes.pop(i)
            st.success("Sheet updated!")
            st.rerun()
        if col2.button("Cancel", key="cancel_" + str(i)):
            st.session_state.pending_writes.pop(i)
            st.rerun()

if prompt := st.chat_input("What do you need today?"):
    st.session_state.messages.append({"role": "user", "content": prompt})
    with st.chat_message("user"):
        st.markdown(prompt)
    with st.chat_message("assistant"):
        with st.spinner("Looking through your files..."):
            response_text, new_writes = ask_claude(prompt, files, sheets_svc, drive_svc)
        st.markdown(response_text)
        st.session_state.pending_writes.extend(new_writes)
        st.session_state.messages.append({"role": "assistant", "content": response_text})
    if new_writes:
        st.rerun()
